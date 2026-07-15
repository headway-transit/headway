"""XLSX alongside CSV (handoff 0017, design point 5).

THE PINNED INVARIANT: for every export surface, the XLSX's cells are
byte-equal to the CSV's values cell-for-cell — same banner lines (CSV
leading lines == XLSX first sheet), same header, same data cells — and
every XLSX cell is a TEXT cell (openpyxl data_type 's'), because an Excel
number cell is an IEEE double and would corrupt an exact NUMERIC figure.
"""

import csv as csv_module
import datetime as dt
import io
import json
from decimal import Decimal

import pytest
from conftest import auth_header
from openpyxl import load_workbook

from headway_api import exports
from headway_api.routers import reports

UTC = dt.timezone.utc


def _csv_rows(data: bytes) -> list[list[str]]:
    return list(csv_module.reader(io.StringIO(data.decode("utf-8"))))


def _xlsx_sheets(data: bytes):
    wb = load_workbook(io.BytesIO(data))
    return wb


def _cell_text(cell) -> str:
    # An empty CSV cell and an empty XLSX cell are both "no content".
    return "" if cell.value is None else cell.value


def assert_csv_xlsx_equal(csv_bytes: bytes, xlsx_bytes: bytes) -> None:
    """The design-point-5 pin, applied to one export: banner lines, header
    and every data cell byte-equal across the two formats; every non-empty
    XLSX cell stored as TEXT."""
    rows = _csv_rows(csv_bytes)
    wb = _xlsx_sheets(xlsx_bytes)

    # Split the CSV into banner lines (single-cell rows before the header)
    # by construction: everything before the first multi-column row.
    banner_lines = []
    i = 0
    while i < len(rows) and len(rows[i]) == 1:
        banner_lines.append(rows[i][0])
        i += 1
    header_and_data = rows[i:]

    if banner_lines:
        assert wb.sheetnames == [
            exports.BANNER_SHEET_TITLE,
            exports.DATA_SHEET_TITLE,
        ]
        banner_sheet = wb[exports.BANNER_SHEET_TITLE]
        xlsx_banner = [
            _cell_text(row[0]) for row in banner_sheet.iter_rows(max_col=1)
        ]
        assert xlsx_banner == banner_lines
    else:
        assert wb.sheetnames == [exports.DATA_SHEET_TITLE]

    data_sheet = wb[exports.DATA_SHEET_TITLE]
    xlsx_rows = [
        [_cell_text(c) for c in row] for row in data_sheet.iter_rows()
    ]
    # Cell-for-cell byte equality (padded: openpyxl trims trailing Nones).
    assert len(xlsx_rows) == len(header_and_data)
    for csv_row, xlsx_row in zip(header_and_data, xlsx_rows):
        padded = xlsx_row + [""] * (len(csv_row) - len(xlsx_row))
        assert padded[: len(csv_row)] == csv_row
    # Every non-empty cell in the workbook is TEXT — never a number cell.
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    assert cell.data_type == "s", (
                        f"non-text cell {cell.coordinate}: "
                        f"{cell.value!r} ({cell.data_type})"
                    )


# ---------------------------------------------------------------------------
# The grid layer itself
# ---------------------------------------------------------------------------


def test_grid_refuses_non_string_cells_and_ragged_rows():
    with pytest.raises(TypeError):
        exports.Grid([], ["a"], [[Decimal("1.5")]])
    with pytest.raises(ValueError):
        exports.Grid([], ["a", "b"], [["only-one"]])


def test_exact_figure_strings_survive_the_xlsx_round_trip():
    """The whole point of text cells: a NUMERIC string that a float would
    mangle comes back byte-identical."""
    value = "12794.920000000000000001"
    grid = exports.Grid(["banner"], ["metric", "value"], [["vrm", value]])
    wb = _xlsx_sheets(exports.grid_to_xlsx(grid))
    cell = wb[exports.DATA_SHEET_TITLE].cell(row=2, column=2)
    assert cell.value == value
    assert cell.data_type == "s"
    assert_csv_xlsx_equal(exports.grid_to_csv(grid), exports.grid_to_xlsx(grid))


# ---------------------------------------------------------------------------
# /metrics/values/export
# ---------------------------------------------------------------------------


def _seed_values(fake_db):
    fake_db.add_metric_value(
        metric="vrm", unit="miles", value=Decimal("12794.92"),
        certification_status="certified",
    )
    fake_db.add_metric_value(
        metric="upt", unit="unlinked_passenger_trips", value=Decimal("238100"),
        detail={"source_mix": {"tides_simulated": 10}},
    )
    fake_db.add_metric_value(
        metric="otp", unit="percent", value=Decimal("54.10"), category="ops",
        calc_name="otp_v0",
    )


def _export(client, fake_db, path, params):
    return client.get(path, params=params, headers=auth_header(fake_db, "vera"))


def test_metrics_values_export_csv_xlsx_byte_equal(client, fake_db):
    _seed_values(fake_db)
    csv_r = _export(client, fake_db, "/metrics/values/export", {"format": "csv"})
    xlsx_r = _export(client, fake_db, "/metrics/values/export", {"format": "xlsx"})
    assert csv_r.status_code == 200 and xlsx_r.status_code == 200
    assert csv_r.headers["content-type"].startswith("text/csv")
    assert xlsx_r.headers["content-type"] == exports.XLSX_MEDIA_TYPE
    assert "attachment" in csv_r.headers["content-disposition"]
    assert_csv_xlsx_equal(csv_r.content, xlsx_r.content)

    rows = _csv_rows(csv_r.content)
    # Banner lines: the preview disclaimer plus (simulated rows exist) the
    # simulated warning.
    assert rows[0] == [exports.METRICS_PREVIEW_DISCLAIMER]
    assert rows[1] == [exports.SIMULATED_BANNER]
    header = rows[2]
    assert header == list(exports.METRICS_VALUES_HEADER)
    by_metric = {r[header.index("metric")]: r for r in rows[3:]}
    # Values verbatim; the simulated row labeled with the web export's words.
    assert by_metric["vrm"][header.index("value")] == "12794.92"
    assert by_metric["upt"][header.index("simulated_data")] == (
        exports.SIMULATED_CELL
    )
    assert by_metric["vrm"][header.index("simulated_data")] == "no"
    assert by_metric["otp"][header.index("category")] == "ops"
    # Provenance travels: every row names its metric_value_id.
    assert all(r[header.index("metric_value_id")] for r in rows[3:])


def test_metrics_values_export_filters_match_the_list(client, fake_db):
    _seed_values(fake_db)
    r = _export(
        client, fake_db, "/metrics/values/export",
        {"format": "csv", "metric": "vrm"},
    )
    rows = _csv_rows(r.content)
    data = [row for row in rows if len(row) > 1][1:]
    assert len(data) == 1 and data[0][0] == "vrm"
    # No simulated row selected → no simulated banner line.
    assert [exports.SIMULATED_BANNER] not in rows


def test_metrics_values_export_bad_format_and_auth(client, fake_db):
    r = _export(client, fake_db, "/metrics/values/export", {"format": "pdf"})
    assert r.status_code == 422
    assert client.get("/metrics/values/export").status_code == 401


# ---------------------------------------------------------------------------
# /reports/mr20/export and /reports/ss50/export (canned packages — the same
# passthrough discipline as test_reports.py)
# ---------------------------------------------------------------------------

MR20_PACKAGE = {
    "form": "MR-20",
    "generator": {"name": "headway_calc.mr20", "version": "0.1.0"},
    "month": "2026-07",
    "period_start": "2026-07-01",
    "period_end": "2026-08-01",
    "period_convention": "half-open [period_start, period_end), UTC",
    "citation": "2025 NTD Monthly Manual pp. 32-33 (tracker)",
    "reportable": False,
    "banner": "NOT REPORTABLE — preview package only.",
    "caveats": [
        {"id": "D2", "status": "open", "text": "Rail passenger-car measure."}
    ],
    "data_points": ["upt", "vrh", "vrm", "voms"],
    "modes": {
        "bus": {
            "upt": {
                "value": "12345",
                "unit": "unlinked_passenger_trips",
                "metric_value_id": "mv-0001",
                "calc_name": "upt_v0",
                "calc_version": "0.1.0",
                "certification_status": "uncertified",
                "flags": ["pre_verification", "uncertified"],
                "coverage": None,
            },
            "vrh": {"value": None, "reason": "No computed row."},
            "vrm": {
                "value": "160835.49",
                "unit": "miles",
                "metric_value_id": "mv-0003",
                "calc_name": "vrm_v0",
                "calc_version": "0.2.0",
                "certification_status": "uncertified",
                "flags": ["pre_verification", "uncertified"],
                "coverage": "0.9932",
            },
            "voms": {"value": None, "reason": "No computed row."},
            "non_reportable_pending_d2": False,
        },
        "rail": {
            "upt": {"value": None, "reason": "No computed row."},
            "vrh": {"value": None, "reason": "No computed row."},
            "vrm": {"value": None, "reason": "No computed row."},
            "voms": {"value": None, "reason": "No computed row."},
            "non_reportable_pending_d2": True,
        },
    },
    "fleet": {
        "upt": {"value": None, "reason": "No computed row."},
        "vrh": {"value": None, "reason": "No computed row."},
        "vrm": {"value": None, "reason": "No computed row."},
        "voms": {
            "value": "42",
            "unit": "vehicles",
            "metric_value_id": "mv-0002",
            "calc_name": "voms_v0",
            "calc_version": "0.1.0",
            "certification_status": "uncertified",
            "flags": ["uncertified", "voms_day_level_proxy"],
            "coverage": None,
        },
    },
}


def test_mr20_export_grid_banner_and_values(client, fake_db, monkeypatch):
    monkeypatch.setattr(
        reports.mr20, "build_mr20_package", lambda conn, month: MR20_PACKAGE
    )
    csv_r = _export(
        client, fake_db, "/reports/mr20/export",
        {"month": "2026-07", "format": "csv"},
    )
    xlsx_r = _export(
        client, fake_db, "/reports/mr20/export",
        {"month": "2026-07", "format": "xlsx"},
    )
    assert csv_r.status_code == 200 and xlsx_r.status_code == 200
    assert_csv_xlsx_equal(csv_r.content, xlsx_r.content)
    rows = _csv_rows(csv_r.content)
    assert rows[0] == [MR20_PACKAGE["banner"]]  # NOT-REPORTABLE leads
    banner_text = "\n".join(r[0] for r in rows if len(r) == 1)
    assert "Caveat [D2, open]" in banner_text
    header = next(r for r in rows if len(r) > 1)
    assert header == list(exports.MR20_HEADER)
    data = [r for r in rows if len(r) > 1][1:]
    # Fleet first, then modes sorted; values verbatim; missing cells carry
    # the package's reason and an empty value.
    assert data[0][0] == "fleet"
    fleet_voms = next(
        r for r in data if r[0] == "fleet" and r[1] == "voms"
    )
    assert fleet_voms[header.index("value")] == "42"
    bus_vrm = next(r for r in data if r[0] == "mode:bus" and r[1] == "vrm")
    assert bus_vrm[header.index("value")] == "160835.49"
    assert bus_vrm[header.index("coverage")] == "0.9932"
    bus_vrh = next(r for r in data if r[0] == "mode:bus" and r[1] == "vrh")
    assert bus_vrh[header.index("value")] == ""
    assert bus_vrh[header.index("missing_reason")] == "No computed row."
    rail_row = next(r for r in data if r[0] == "mode:rail")
    assert rail_row[header.index("non_reportable_pending_d2")] == "yes"


def test_mr20_export_bad_month_is_422(client, fake_db):
    r = _export(
        client, fake_db, "/reports/mr20/export",
        {"month": "July", "format": "csv"},
    )
    assert r.status_code == 422


SS50_PACKAGE = {
    "form": "S&S-50",
    "generator": {"name": "headway_calc.ss50", "version": "0.1.0"},
    "month": "2026-06",
    "period_start": "2026-06-01",
    "period_end": "2026-07-01",
    "period_convention": (
        "half-open [period_start, period_end), UTC, on occurred_at"
    ),
    "due_date": "2026-07-31",
    "reportable": False,
    "banner": "NOT REPORTABLE — preview package only.",
    "citations": [
        {"id": "ss50_scope", "text": "p. 3 scope.", "source": "tracker"}
    ],
    "caveats": [{"id": "tos_attribution", "text": "TOS is manual entry."}],
    "operated_modes": ["bus"],
    "cells": [
        {
            "mode": "bus",
            "type_of_service": "DO",
            "zero_event": False,
            "counts": {
                "injury_events": {
                    "count": 2, "people_injured": 3,
                    "event_ids": ["ev-1", "ev-2"],
                },
                "non_major_fires": {"count": 1, "event_ids": ["ev-3"]},
                "assaults_on_worker": {
                    "count": 1, "without_injury": 1, "event_ids": ["ev-4"],
                },
            },
        },
        {
            "mode": "rail",
            "type_of_service": "unknown",
            "zero_event": True,
            "counts": {
                "injury_events": {
                    "count": 0, "people_injured": 0, "event_ids": [],
                },
                "non_major_fires": {"count": 0, "event_ids": []},
                "assaults_on_worker": {
                    "count": 0, "without_injury": 0, "event_ids": [],
                },
            },
        },
    ],
    "excluded": {
        "major_event_ids": ["ev-9"],
        "not_reportable_event_ids": [],
        "superseded_event_ids": [],
        "unclassified_event_ids": ["ev-8"],
    },
}


def test_ss50_export_grid_banner_and_values(client, fake_db, monkeypatch):
    monkeypatch.setattr(
        reports.ss50, "build_ss50_package", lambda conn, month: SS50_PACKAGE
    )
    csv_r = _export(
        client, fake_db, "/reports/ss50/export",
        {"month": "2026-06", "format": "csv"},
    )
    xlsx_r = _export(
        client, fake_db, "/reports/ss50/export",
        {"month": "2026-06", "format": "xlsx"},
    )
    assert csv_r.status_code == 200 and xlsx_r.status_code == 200
    assert_csv_xlsx_equal(csv_r.content, xlsx_r.content)
    rows = _csv_rows(csv_r.content)
    assert rows[0] == [SS50_PACKAGE["banner"]]
    banner_text = "\n".join(r[0] for r in rows if len(r) == 1)
    assert "Due date: 2026-07-31" in banner_text
    assert "ev-9" in banner_text  # excluded major named
    assert "ev-8" in banner_text  # unclassified named
    header = next(r for r in rows if len(r) > 1)
    assert header == list(exports.SS50_HEADER)
    data = [r for r in rows if len(r) > 1][1:]
    bus = next(r for r in data if r[0] == "bus")
    assert bus[header.index("injury_events")] == "2"
    assert bus[header.index("people_injured")] == "3"
    assert bus[header.index("injury_event_ids")] == "ev-1; ev-2"
    assert bus[header.index("zero_event")] == "no"
    rail = next(r for r in data if r[0] == "rail")
    assert rail[header.index("zero_event")] == "yes"  # explicit zero row


# ---------------------------------------------------------------------------
# /sampling/plans/{id}/worksheet
# ---------------------------------------------------------------------------


def test_sampling_worksheet_export(client, fake_db):
    plan = fake_db.add_sampling_plan(required_per_period=2, required_annual=4)
    fake_db.add_sampling_draw(
        plan["plan_id"], period_label="2026-Q1",
        service_units=["u1", "u2", "u3"], selected_units=["u1", "u2"],
    )
    fake_db.add_sampling_draw(
        plan["plan_id"], period_label="2026-Q2",
        service_units=["u4", "u5", "u6"], selected_units=["u4", "u5"],
        oversample_units=1,
    )
    fake_db.add_sampling_measurement(plan["plan_id"], "u1")
    fake_db.add_sampling_measurement(plan["plan_id"], "u4")

    csv_r = _export(
        client, fake_db,
        f"/sampling/plans/{plan['plan_id']}/worksheet", {"format": "csv"},
    )
    xlsx_r = _export(
        client, fake_db,
        f"/sampling/plans/{plan['plan_id']}/worksheet", {"format": "xlsx"},
    )
    assert csv_r.status_code == 200 and xlsx_r.status_code == 200
    assert_csv_xlsx_equal(csv_r.content, xlsx_r.content)
    rows = _csv_rows(csv_r.content)
    banner_text = "\n".join(r[0] for r in rows if len(r) == 1)
    assert "UNDERSAMPLED" in banner_text  # 2 of 4 measured
    assert "2 of 4 required" in banner_text
    header = next(r for r in rows if len(r) > 1)
    assert header == list(exports.SAMPLING_WORKSHEET_HEADER)
    data = [r for r in rows if len(r) > 1][1:]
    assert [
        (r[0], r[1], r[2], r[3]) for r in data
    ] == [
        ("2026-Q1", "u1", "yes", "0"),
        ("2026-Q1", "u2", "no", "0"),
        ("2026-Q2", "u4", "yes", "1"),
        ("2026-Q2", "u5", "no", "1"),
    ]


def test_sampling_worksheet_estimate_ready_state(client, fake_db):
    plan = fake_db.add_sampling_plan(required_per_period=1, required_annual=1)
    fake_db.add_sampling_draw(
        plan["plan_id"], period_label="2026-Q1",
        service_units=["u1"], selected_units=["u1"],
    )
    fake_db.add_sampling_measurement(plan["plan_id"], "u1")
    r = _export(
        client, fake_db,
        f"/sampling/plans/{plan['plan_id']}/worksheet", {"format": "csv"},
    )
    banner_text = "\n".join(
        row[0] for row in _csv_rows(r.content) if len(row) == 1
    )
    assert "estimate-ready" in banner_text


def test_sampling_worksheet_unknown_plan_404(client, fake_db):
    r = _export(
        client, fake_db,
        "/sampling/plans/00000000-0000-0000-0000-000000000000/worksheet",
        {"format": "csv"},
    )
    assert r.status_code == 404


def test_superseded_measurement_does_not_count_as_measured(client, fake_db):
    """The worksheet uses ACTIVE measurements only — a superseded row's
    unit stays 'measured' only through its replacement."""
    plan = fake_db.add_sampling_plan(required_per_period=1, required_annual=1)
    fake_db.add_sampling_draw(
        plan["plan_id"], period_label="2026-Q1",
        service_units=["u1"], selected_units=["u1"],
    )
    m = fake_db.add_sampling_measurement(plan["plan_id"], "u1")
    m["superseded_by"] = "some-replacement-id"  # no active replacement seeded
    r = _export(
        client, fake_db,
        f"/sampling/plans/{plan['plan_id']}/worksheet", {"format": "csv"},
    )
    rows = _csv_rows(r.content)
    data = [row for row in rows if len(row) > 1][1:]
    assert data == [["2026-Q1", "u1", "no", "0"]]
