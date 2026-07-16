"""GET /reports/agency-workbook (handoff 0020): the monthly agency workbook
— banner sheet first, stated-absent cells, a VISIBLE provenance column,
migration-0024 ops badging, verbatim year-over-year columns, and the
multi-sheet CSV/XLSX byte-equality invariant."""

import csv as csv_module
import datetime as dt
import io
from decimal import Decimal

from conftest import auth_header
from openpyxl import load_workbook

from headway_api import exports

UTC = dt.timezone.utc

JULY = (dt.date(2026, 7, 1), dt.date(2026, 8, 1))
PRIOR_JULY = (dt.date(2025, 7, 1), dt.date(2025, 8, 1))


def _seed(fake_db):
    """A month with: agency + one mode UPT, day-type averages (typical +
    one atypical split), days operated, VOMS, one ops row — and a
    prior-year UPT row for the YoY column."""
    ids = {}
    ids["upt"] = fake_db.add_metric_value(
        metric="upt", unit="unlinked_passenger_trips",
        value=Decimal("123456"), scope="agency",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="upt_v0", calc_version="0.2.0",
        detail={"source_mix": {"tides_simulated": 10}},
    )["metric_value_id"]
    ids["upt_bus"] = fake_db.add_metric_value(
        metric="upt", unit="unlinked_passenger_trips",
        value=Decimal("100000"), scope="mode:bus",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="upt_v0", calc_version="0.2.0",
        detail={"source_mix": {"tides_simulated": 8}},
    )["metric_value_id"]
    ids["avg_weekday"] = fake_db.add_metric_value(
        metric="upt_avg", unit="unlinked_passenger_trips_per_day",
        value=Decimal("4567.89"), scope="daytype:weekday",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="daytype_upt_avg_v0", calc_version="0.1.0",
        detail={
            "day_type": "weekday", "split": "typical",
            "days_operated": 21, "atypical_flags_declared": False,
            "source_mix": {"tides_simulated": 100},
        },
    )["metric_value_id"]
    ids["avg_saturday_atypical"] = fake_db.add_metric_value(
        metric="upt_avg", unit="unlinked_passenger_trips_per_day",
        value=Decimal("999.99"), scope="daytype:saturday:atypical",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="daytype_upt_avg_v0", calc_version="0.1.0",
        detail={
            "day_type": "saturday", "split": "atypical",
            "days_operated": 1, "atypical_flags_declared": True,
            "source_mix": {"tides_simulated": 5},
        },
    )["metric_value_id"]
    ids["days_weekday"] = fake_db.add_metric_value(
        metric="days_operated", unit="days",
        value=Decimal("21"), scope="daytype:weekday",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="daytype_days_operated_v0", calc_version="0.1.0",
        detail={
            "day_type": "weekday",
            "unobserved_dates": ["2026-07-30", "2026-07-31"],
        },
    )["metric_value_id"]
    ids["voms"] = fake_db.add_metric_value(
        metric="voms", unit="vehicles",
        value=Decimal("42"), scope="agency",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="voms_v0", calc_version="0.1.0", detail={},
    )["metric_value_id"]
    ids["otp"] = fake_db.add_metric_value(
        metric="otp", unit="share_on_time",
        value=Decimal("0.5410"), scope="agency",
        period_start=JULY[0], period_end=JULY[1],
        calc_name="otp_v0", calc_version="0.1.0",
        category="ops", detail={},
    )["metric_value_id"]
    ids["upt_prior"] = fake_db.add_metric_value(
        metric="upt", unit="unlinked_passenger_trips",
        value=Decimal("111111"), scope="agency",
        period_start=PRIOR_JULY[0], period_end=PRIOR_JULY[1],
        calc_name="upt_v0", calc_version="0.1.0",
        detail={"source_mix": {"tides_simulated": 9}},
    )["metric_value_id"]
    return ids


def _fetch(client, fake_db, fmt):
    return client.get(
        "/reports/agency-workbook",
        params={"month": "2026-07", "format": fmt},
        headers=auth_header(fake_db, "vera"),
    )


def _csv_sections(data: bytes):
    """Split the multi-sheet CSV into (banner_lines, {title: rows})."""
    rows = list(csv_module.reader(io.StringIO(data.decode("utf-8"))))
    banner, sections, current = [], {}, None
    for row in rows:
        if len(row) == 1 and row[0].startswith(exports.SHEET_MARKER_PREFIX):
            current = row[0][len(exports.SHEET_MARKER_PREFIX):]
            sections[current] = []
        elif current is None:
            assert len(row) == 1, row  # banner lines are single-cell
            banner.append(row[0])
        else:
            sections[current].append(row)
    return banner, sections


def test_workbook_csv_and_xlsx_are_cell_for_cell_equal(client, fake_db):
    _seed(fake_db)
    csv_bytes = _fetch(client, fake_db, "csv").content
    xlsx_bytes = _fetch(client, fake_db, "xlsx").content
    banner, sections = _csv_sections(csv_bytes)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == [exports.BANNER_SHEET_TITLE] + list(sections)

    xlsx_banner = [
        "" if row[0].value is None else row[0].value
        for row in wb[exports.BANNER_SHEET_TITLE].iter_rows(max_col=1)
    ]
    assert xlsx_banner == banner

    for title, csv_rows in sections.items():
        sheet = wb[title]
        xlsx_rows = [
            ["" if c.value is None else c.value for c in row]
            for row in sheet.iter_rows()
        ]
        assert len(xlsx_rows) == len(csv_rows), title
        for csv_row, xlsx_row in zip(csv_rows, xlsx_rows):
            padded = xlsx_row + [""] * (len(csv_row) - len(xlsx_row))
            assert padded[: len(csv_row)] == csv_row
    # Every non-empty cell is TEXT — a figure never becomes an IEEE double.
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    assert cell.data_type == "s", cell.coordinate


def test_workbook_banner_states_scope_and_absences(client, fake_db):
    _seed(fake_db)
    banner, _sections = _csv_sections(_fetch(client, fake_db, "csv").content)
    joined = "\n".join(banner)
    assert banner[0] == exports.WORKBOOK_BANNER  # NOT REPORTABLE leads
    assert "missed-trip accounting" in joined  # stated future increment
    assert "VAMS" in joined  # stated absent
    assert exports.WORKBOOK_OPS_BADGE in joined  # the 0024 wall, stated
    assert exports.SIMULATED_BANNER in joined  # simulated rows present
    assert "daytype_v0 0.1.0" in joined  # classification basis named
    assert "GET /metrics/compare" in joined  # no derived deltas — stated


def test_workbook_rows_provenance_and_stated_absences(client, fake_db):
    ids = _seed(fake_db)
    _banner, sections = _csv_sections(_fetch(client, fake_db, "csv").content)
    ridership = sections[exports.WORKBOOK_SHEET_RIDERSHIP]
    header = ridership[0]
    assert header == list(exports.WORKBOOK_HEADER)
    idx = {name: header.index(name) for name in header}
    by_measure_scope = {
        (r[idx["measure"]], r[idx["scope"]]): r for r in ridership[1:]
    }

    # Present cell: verbatim value + visible provenance + YoY verbatim.
    upt = by_measure_scope[
        ("Unlinked Passenger Trips (month total)", "all modes (agency)")
    ]
    assert upt[idx["value"]] == "123456"
    assert upt[idx["provenance_metric_value_id"]] == ids["upt"]
    assert upt[idx["prior_year_value"]] == "111111"
    assert upt[idx["prior_year_provenance"]] == ids["upt_prior"]
    assert upt[idx["simulated_data"]] == exports.SIMULATED_CELL

    # Day-type average: split + all-typical statement from the detail.
    avg = by_measure_scope[
        ("Average weekday UPT (typical days)", "all modes (agency)")
    ]
    assert avg[idx["value"]] == "4567.89"
    assert avg[idx["provenance_metric_value_id"]] == ids["avg_weekday"]
    assert "21 operated" in avg[idx["note"]]
    assert "every day is typical" in avg[idx["note"]]

    # The declared-atypical split gets its own row.
    atypical = by_measure_scope[
        (
            "Average saturday UPT (agency-declared atypical days)",
            "all modes (agency)",
        )
    ]
    assert atypical[idx["value"]] == "999.99"

    # Absent cell: STATED, empty provenance, never zero-filled.
    sunday_avg = by_measure_scope[
        ("Average sunday UPT (typical days)", "all modes (agency)")
    ]
    assert sunday_avg[idx["value"]] == exports.WORKBOOK_ABSENT_CELL
    assert sunday_avg[idx["provenance_metric_value_id"]] == ""
    assert "never invented" in sunday_avg[idx["note"]]

    # Days operated: observed-lower-bound note from the persisted detail.
    days = by_measure_scope[
        ("Days operated (weekday schedule)", "all modes (agency)")
    ]
    assert days[idx["value"]] == "21"
    assert "observed lower bound: 2 date(s)" in days[idx["note"]]

    # The honest-scope rows exist and are stated absent.
    missed = by_measure_scope[
        ("Missed trips / trips not operated", "all modes (agency)")
    ]
    assert missed[idx["value"]] == exports.WORKBOOK_ABSENT_CELL
    assert "schedule-vs-operated reconciliation" in missed[idx["note"]]


def test_workbook_operations_sheet_badges_ops_rows(client, fake_db):
    ids = _seed(fake_db)
    _banner, sections = _csv_sections(_fetch(client, fake_db, "csv").content)
    operations = sections[exports.WORKBOOK_SHEET_OPERATIONS]
    header = operations[0]
    idx = {name: header.index(name) for name in header}
    by_measure = {
        (r[idx["measure"]], r[idx["scope"]]): r for r in operations[1:]
    }

    voms = by_measure[
        ("Vehicles Operated in Maximum Service (VOMS)", "all modes (agency)")
    ]
    assert voms[idx["value"]] == "42"
    assert voms[idx["category"]] == "ntd"
    # The mode without a VOMS row is stated absent, never zero-filled.
    voms_bus = by_measure[
        ("Vehicles Operated in Maximum Service (VOMS)", "mode:bus")
    ]
    assert voms_bus[idx["value"]] == exports.WORKBOOK_ABSENT_CELL

    otp = by_measure[
        (
            "On-time performance (share of observed passages on time)",
            "all modes (agency)",
        )
    ]
    assert otp[idx["value"]] == "0.5410"
    assert otp[idx["category"]] == "ops"
    assert otp[idx["note"]] == exports.WORKBOOK_OPS_BADGE
    assert otp[idx["provenance_metric_value_id"]] == ids["otp"]
    # The database CHECK (0024) makes a certified ops row unrepresentable;
    # the workbook serves its status verbatim.
    assert otp[idx["certification_status"]] == "uncertified"

    vams = by_measure[
        (
            "Vehicles Available for Maximum Service (VAMS)",
            "all modes (agency)",
        )
    ]
    assert vams[idx["value"]] == exports.WORKBOOK_ABSENT_CELL
    assert "fleet inventory" in vams[idx["note"]]


def test_workbook_auth_and_month_validation(client, fake_db):
    assert (
        client.get(
            "/reports/agency-workbook", params={"month": "2026-07"}
        ).status_code
        == 401
    )
    bad = client.get(
        "/reports/agency-workbook",
        params={"month": "July 2026"},
        headers=auth_header(fake_db, "vera"),
    )
    assert bad.status_code == 422
    assert "YYYY-MM" in bad.json()["detail"]
    bad_format = client.get(
        "/reports/agency-workbook",
        params={"month": "2026-07", "format": "pdf"},
        headers=auth_header(fake_db, "vera"),
    )
    assert bad_format.status_code == 422


def test_certificate_lines_join_the_banner_when_figures_are_certified(
    client, fake_db
):
    _seed(fake_db)
    mv = fake_db.add_metric_value(
        metric="upt", unit="unlinked_passenger_trips",
        value=Decimal("5"), scope="mode:certified-demo",
        period_start=JULY[0], period_end=JULY[1],
        certification_status="certified", detail={},
    )
    fake_db.certifications.append(
        {
            "certification_id": "cert-1",
            "metric_value_ids": [mv["metric_value_id"]],
            "certified_by": "cora",
            "certified_at": dt.datetime(2026, 8, 2, 9, 0, tzinfo=UTC),
            "attestation": "attested",
            "canonical_document": None,
            "signature": None,
            "key_fingerprint": None,
        }
    )
    banner, _sections = _csv_sections(_fetch(client, fake_db, "csv").content)
    joined = "\n".join(banner)
    assert "Certification cert-1" in joined
    assert "honest history, never backfilled" in joined
