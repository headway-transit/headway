"""The certifier digital signature (handoff 0019, design point B): canonical
document, Ed25519 signing, verification, tamper-evidence, honest legacy
NULLs, and the propagation surfaces (public fingerprint, export certificate
block).
"""

from __future__ import annotations

import base64
import datetime as dt
import json

import pytest
from conftest import (
    TEST_SIGNING_SEED_HEX,
    UTC,
    auth_header,
)

from headway_api import signing


def _certify(client, fake_db, ids, **overrides):
    payload = {
        "metric_value_ids": ids,
        "attestation": "I certify these figures are accurate.",
        "signer_full_name": "Cora Certifier",
        "signer_title": "Chief Executive Officer",
    }
    payload.update(overrides)
    return client.post(
        "/certifications", json=payload, headers=auth_header(fake_db, "cora")
    )


# --- canonicalization ------------------------------------------------------------


def test_canonical_bytes_are_deterministic_and_key_sorted():
    a = signing.canonical_bytes({"b": 1, "a": {"z": "ü", "y": [1, 2]}})
    b = signing.canonical_bytes({"a": {"y": [1, 2], "z": "ü"}, "b": 1})
    assert a == b
    assert a == b'{"a":{"y":[1,2],"z":"\\u00fc"},"b":1}'


def test_canonical_bytes_refuse_nan():
    with pytest.raises(ValueError):
        signing.canonical_bytes({"x": float("nan")})


def test_receipt_sha256_recomputable():
    doc = {"metric": "upt", "value": "100"}
    assert signing.receipt_sha256(doc) == signing.receipt_sha256(dict(doc))
    assert len(signing.receipt_sha256(doc)) == 64


# --- key handling ------------------------------------------------------------------


def test_load_signer_from_env_hex_and_fingerprint_shape():
    signer = signing.load_signer({signing.ENV_KEY: TEST_SIGNING_SEED_HEX})
    assert signer.key_fingerprint.startswith("ed25519:")
    assert len(signer.key_fingerprint) == len("ed25519:") + 64
    # Same seed → same fingerprint (deterministic identity).
    again = signing.load_signer({signing.ENV_KEY: TEST_SIGNING_SEED_HEX})
    assert again.key_fingerprint == signer.key_fingerprint


def test_load_signer_refuses_missing_or_malformed_key():
    with pytest.raises(signing.SigningKeyUnavailable):
        signing.load_signer({})
    with pytest.raises(signing.SigningKeyMalformed):
        signing.load_signer({signing.ENV_KEY: "too-short"})
    with pytest.raises(signing.SigningKeyMalformed):
        signing.load_signer({signing.ENV_KEY: "zz" * 32})


def test_load_signer_generates_key_file_at_first_use(tmp_path):
    key_file = tmp_path / "signing.key"
    signer = signing.load_signer(
        {signing.ENV_KEY_FILE: str(key_file)}
    )
    assert key_file.exists()
    assert oct(key_file.stat().st_mode & 0o777) == "0o600"
    # Second load reads the SAME key back.
    again = signing.load_signer({signing.ENV_KEY_FILE: str(key_file)})
    assert again.key_fingerprint == signer.key_fingerprint


# --- the signing certification --------------------------------------------------


def test_certify_signs_canonical_document_with_receipt_hashes(
    client, fake_db, test_signer
):
    mv = fake_db.add_metric_value(
        metric="upt",
        unit="unlinked_passenger_trips",
        calc_name="upt_v0",
        calc_version="0.2.0",
        detail={
            "factor_applied": "10.000000",
            "attestation": {
                "attestation_id": "att-42",
                "statistician_name": "Dr. R. Fisher",
            },
        },
    )
    r = _certify(client, fake_db, [mv["metric_value_id"]])
    assert r.status_code == 201
    body = r.json()
    assert body["signer_full_name"] == "Cora Certifier"
    assert body["signer_title"] == "Chief Executive Officer"
    assert body["key_fingerprint"] == test_signer.key_fingerprint
    assert body["algorithm"] == "ed25519"

    document = json.loads(body["canonical_document"])
    # The stored text is EXACTLY the canonical bytes of its parsed form.
    assert (
        signing.canonical_bytes(document).decode("utf-8")
        == body["canonical_document"]
    )
    assert document["document_type"] == "headway-certification"
    assert document["certification_id"] == body["certification_id"]
    assert document["certifier"] == {
        "username": "cora",
        "role": "certifying_official",
        "typed_full_name": "Cora Certifier",
        "typed_title": "Chief Executive Officer",
    }
    assert document["intent_statement"]  # ESIGN-style statement travels
    assert "not a personal" in document["scope_statement"]  # honest scope
    # Figures with independently recomputable receipt hashes.
    (figure,) = document["figures"]
    assert figure["metric_value_id"] == mv["metric_value_id"]
    assert figure["value"] == str(mv["value"])
    claimed = figure.pop("receipt_sha256")
    assert claimed == signing.receipt_sha256(figure)
    # Statistician attestations the figures carry are acknowledged ON the
    # certificate.
    assert document["statistician_attestations"] == [
        {"attestation_id": "att-42", "statistician_name": "Dr. R. Fisher"}
    ]
    # The signature verifies over the stored bytes.
    assert test_signer.verify(body["canonical_document"], body["signature"])
    # Stored row carries the trio; audit carries the fingerprint.
    (cert,) = fake_db.certifications
    assert cert["canonical_document"] == body["canonical_document"]
    assert cert["signature"] == body["signature"]
    assert cert["key_fingerprint"] == test_signer.key_fingerprint
    (audit,) = [e for e in fake_db.audit_events if e["action"] == "certify"]
    detail = json.loads(audit["detail"])
    assert detail["key_fingerprint"] == test_signer.key_fingerprint
    assert detail["signer_full_name"] == "Cora Certifier"


def test_certify_requires_typed_name_and_title(client, fake_db):
    mv = fake_db.add_metric_value()
    r = client.post(
        "/certifications",
        json={
            "metric_value_ids": [mv["metric_value_id"]],
            "attestation": "ok",
        },
        headers=auth_header(fake_db, "cora"),
    )
    assert r.status_code == 422
    assert fake_db.certifications == []


def test_certify_refuses_503_without_signing_key_nothing_written(
    fake_db, settings, monkeypatch
):
    """A certification is never silently recorded unsigned."""
    from fastapi.testclient import TestClient

    from headway_api.app import create_app

    monkeypatch.delenv(signing.ENV_KEY, raising=False)
    monkeypatch.delenv(signing.ENV_KEY_FILE, raising=False)
    app = create_app(settings=settings, db=fake_db)
    mv = fake_db.add_metric_value()
    with TestClient(app) as client:
        r = _certify(client, fake_db, [mv["metric_value_id"]])
    assert r.status_code == 503
    assert "signing key" in r.json()["detail"]
    assert fake_db.certifications == []
    assert mv["certification_status"] == "uncertified"


# --- the record and certificate views ---------------------------------------------


def test_list_and_certificate_view(client, fake_db, test_signer):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    r = client.get("/certifications", headers=auth_header(fake_db, "vera"))
    assert r.status_code == 200
    (record,) = r.json()
    assert record["certification_id"] == created["certification_id"]
    assert record["signed"] is True
    assert record["signer_full_name"] == "Cora Certifier"
    assert record["key_fingerprint"] == test_signer.key_fingerprint

    r = client.get(
        f"/certifications/{created['certification_id']}",
        headers=auth_header(fake_db, "vera"),
    )
    assert r.status_code == 200
    certificate = r.json()
    assert certificate["document"]["certification_id"] == created[
        "certification_id"
    ]
    assert certificate["verification"]["verdict"] == "verified"
    assert certificate["verification"]["verified"] is True


def test_legacy_unsigned_certification_reads_honestly(client, fake_db):
    fake_db.certifications.append(
        {
            "certification_id": "11111111-1111-1111-1111-111111111111",
            "metric_value_ids": [],
            "certified_by": "cora",
            "certified_at": dt.datetime(2026, 7, 1, tzinfo=UTC),
            "attestation": "pre-signature certification",
            "canonical_document": None,
            "signature": None,
            "key_fingerprint": None,
        }
    )
    r = client.get(
        "/certifications/11111111-1111-1111-1111-111111111111/verify",
        headers=auth_header(fake_db, "vera"),
    )
    assert r.status_code == 200
    body = r.json()
    assert body["signed"] is False
    assert body["verified"] is None
    assert body["verdict"] == "unsigned_legacy"
    assert "never backfilled" in body["message"]


# --- verification and tamper-evidence ----------------------------------------------


def test_verify_endpoint_verifies_untampered_record(client, fake_db):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    r = client.get(
        f"/certifications/{created['certification_id']}/verify",
        headers=auth_header(fake_db, "vera"),
    )
    assert r.status_code == 200
    assert r.json()["verdict"] == "verified"


def test_tampered_canonical_document_fails_loudly(client, fake_db):
    """THE tamper test (handoff 0019): mutate the stored document — the
    live analogue SQL-mutates the row in a scratch database (trigger
    disabled, simulating out-of-band access); the verify endpoint must
    scream."""
    mv = fake_db.add_metric_value(value="1234.567")
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    (cert,) = fake_db.certifications
    # An attacker inflates the certified figure inside the stored document.
    cert["canonical_document"] = cert["canonical_document"].replace(
        "1234.567", "9234.567"
    )
    r = client.get(
        f"/certifications/{created['certification_id']}/verify",
        headers=auth_header(fake_db, "vera"),
    )
    assert r.status_code == 200
    body = r.json()
    assert body["verified"] is False
    assert body["verdict"] == "failed"
    assert "VERIFICATION FAILED" in body["message"]
    assert "tampered" in body["message"]


def test_tampered_signature_fails_loudly(client, fake_db):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    (cert,) = fake_db.certifications
    raw = bytearray(base64.b64decode(cert["signature"]))
    raw[0] ^= 0xFF
    cert["signature"] = base64.b64encode(bytes(raw)).decode("ascii")
    r = client.get(
        f"/certifications/{created['certification_id']}/verify",
        headers=auth_header(fake_db, "vera"),
    )
    assert r.json()["verdict"] == "failed"


def test_swapped_document_from_another_record_fails_binding(
    client, fake_db
):
    """A cryptographically valid (document, signature) pair copied from a
    DIFFERENT certification must fail: the document is bound to its row."""
    mv1 = fake_db.add_metric_value()
    mv2 = fake_db.add_metric_value()
    c1 = _certify(client, fake_db, [mv1["metric_value_id"]]).json()
    c2 = _certify(client, fake_db, [mv2["metric_value_id"]]).json()
    cert2 = next(
        c
        for c in fake_db.certifications
        if c["certification_id"] == c2["certification_id"]
    )
    # Swap in certification 1's (valid!) document + signature.
    cert2["canonical_document"] = c1["canonical_document"]
    cert2["signature"] = c1["signature"]
    r = client.get(
        f"/certifications/{c2['certification_id']}/verify",
        headers=auth_header(fake_db, "vera"),
    )
    body = r.json()
    assert body["verified"] is False
    assert body["verdict"] == "failed"
    assert "not bound to this certification" in body["message"]


def test_key_mismatch_reports_honestly_not_as_tampering(
    client, fake_db, app
):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    # The installation rotates its key.
    app.state.signer = signing.load_signer({signing.ENV_KEY: "cd" * 32})
    r = client.get(
        f"/certifications/{created['certification_id']}/verify",
        headers=auth_header(fake_db, "vera"),
    )
    body = r.json()
    assert body["verdict"] == "key_mismatch"
    assert body["verified"] is False
    assert "UNVERIFIED, not as proof of tampering" in body["message"]


# --- public propagation (design point 7) --------------------------------------------


def test_public_certified_endpoint_serves_fingerprint_no_identity(
    client, fake_db, test_signer
):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    r = client.get("/public/metrics/certified")  # no auth — public
    assert r.status_code == 200
    (row,) = r.json()
    assert row["certification"]["certification_id"] == created[
        "certification_id"
    ]
    assert row["certification"]["key_fingerprint"] == test_signer.key_fingerprint
    # No certifier identity anywhere in the public payload.
    assert "Cora" not in r.text
    assert "cora" not in r.text


def test_public_verify_endpoint_no_auth_no_identity(client, fake_db):
    mv = fake_db.add_metric_value()
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    r = client.get(
        f"/public/certifications/{created['certification_id']}/verify"
    )
    assert r.status_code == 200
    body = r.json()
    assert body["verdict"] == "verified"
    assert "Cora" not in r.text and "cora" not in r.text
    # Tampered → the public check screams too.
    (cert,) = fake_db.certifications
    cert["canonical_document"] += " "
    r = client.get(
        f"/public/certifications/{created['certification_id']}/verify"
    )
    assert r.json()["verdict"] == "failed"


def test_public_verify_unknown_id_404(client, fake_db):
    r = client.get(
        "/public/certifications/00000000-0000-0000-0000-000000000000/verify"
    )
    assert r.status_code == 404


# --- export certificate block (design point 7) --------------------------------------

#: Minimal canned packages (the reports endpoints are passthroughs; the
#: package builders need a psycopg cursor the FakeConn does not model —
#: the test_reports.py monkeypatch precedent).
_CANNED_MR20 = {
    "form": "MR-20",
    "generator": {"name": "headway_calc.mr20", "version": "0.1.0"},
    "month": "2026-06",
    "period_start": "2026-06-01",
    "period_end": "2026-07-01",
    "period_convention": "half-open [period_start, period_end), UTC",
    "citation": "2026 NTD Policy Manual (Full Reporting).",
    "reportable": False,
    "banner": "NOT REPORTABLE — preview package only.",
    "caveats": [],
    "modes": {},
    "fleet": {},
}

_CANNED_SS50 = {
    "form": "S&S-50",
    "generator": {"name": "headway_calc.ss50", "version": "0.1.0"},
    "month": "2026-06",
    "period_start": "2026-06-01",
    "period_end": "2026-07-01",
    "period_convention": "half-open [period_start, period_end), UTC",
    "due_date": "2026-07-30",
    "banner": "NOT REPORTABLE — preview package only.",
    "citations": [],
    "caveats": [],
    "excluded": {},
    "cells": [],
}


def _patch_packages(monkeypatch):
    from headway_api.routers import reports

    monkeypatch.setattr(
        reports.mr20, "build_mr20_package", lambda conn, month: _CANNED_MR20
    )
    monkeypatch.setattr(
        reports.ss50, "build_ss50_package", lambda conn, month: _CANNED_SS50
    )


def _mr20_month_figure(fake_db, **overrides):
    return fake_db.add_metric_value(
        metric="upt",
        unit="unlinked_passenger_trips",
        calc_name="upt_v0",
        calc_version="0.2.0",
        period_start=dt.date(2026, 6, 1),
        period_end=dt.date(2026, 7, 1),
        value="100",
        **overrides,
    )


def test_mr20_export_gains_certificate_block_when_period_certified(
    client, fake_db, test_signer, monkeypatch
):
    _patch_packages(monkeypatch)
    mv = _mr20_month_figure(fake_db)
    before = client.get(
        "/reports/mr20/export?month=2026-06&format=csv",
        headers=auth_header(fake_db, "petra"),
    )
    assert before.status_code == 200
    assert "Certification " not in before.text

    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    after = client.get(
        "/reports/mr20/export?month=2026-06&format=csv",
        headers=auth_header(fake_db, "petra"),
    )
    assert after.status_code == 200
    assert created["certification_id"] in after.text
    assert "Cora Certifier, Chief Executive Officer" in after.text
    assert test_signer.key_fingerprint in after.text
    assert "/public/certifications/{certification_id}/verify" in after.text

    # The XLSX carries the same lines on the "Read first" sheet.
    import io

    from openpyxl import load_workbook

    xlsx = client.get(
        "/reports/mr20/export?month=2026-06&format=xlsx",
        headers=auth_header(fake_db, "petra"),
    )
    wb = load_workbook(io.BytesIO(xlsx.content))
    assert wb.sheetnames[0] == "Read first"
    lines = [
        str(row[0].value)
        for row in wb["Read first"].iter_rows()
        if row[0].value is not None
    ]
    assert any(test_signer.key_fingerprint in line for line in lines)
    assert any(created["certification_id"] in line for line in lines)


def test_ss50_export_gains_certificate_block_too(
    client, fake_db, test_signer, monkeypatch
):
    _patch_packages(monkeypatch)
    mv = _mr20_month_figure(fake_db)
    created = _certify(client, fake_db, [mv["metric_value_id"]]).json()
    r = client.get(
        "/reports/ss50/export?month=2026-06&format=csv",
        headers=auth_header(fake_db, "petra"),
    )
    assert r.status_code == 200
    assert created["certification_id"] in r.text
    assert test_signer.key_fingerprint in r.text


def test_export_names_legacy_certifications_honestly(
    client, fake_db, monkeypatch
):
    _patch_packages(monkeypatch)
    mv = _mr20_month_figure(fake_db, certification_status="certified")
    fake_db.certifications.append(
        {
            "certification_id": "22222222-2222-2222-2222-222222222222",
            "metric_value_ids": [mv["metric_value_id"]],
            "certified_by": "cora",
            "certified_at": dt.datetime(2026, 7, 1, tzinfo=UTC),
            "attestation": "legacy",
            "canonical_document": None,
            "signature": None,
            "key_fingerprint": None,
        }
    )
    r = client.get(
        "/reports/mr20/export?month=2026-06&format=csv",
        headers=auth_header(fake_db, "petra"),
    )
    assert "recorded before digital signatures existed" in r.text
