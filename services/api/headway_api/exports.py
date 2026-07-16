"""Tabular exports — CSV and XLSX from ONE row assembly (handoff 0017,
design point 5).

THE INVARIANT (pinned by test): for every export surface, the CSV and the
XLSX are generated from the same assembled grid, and every XLSX data cell is
written as a TEXT cell holding the byte-identical string the CSV holds.
Figures reach this module as the exact strings the API already serves
(NUMERIC → string; never float) and leave it untouched — this module applies
formatting (CSV quoting, workbook structure) and NEVER arithmetic, rounding,
or type coercion on a value. Text cells are the point: an Excel 'number'
cell is an IEEE double, which would silently corrupt an exact NUMERIC
figure; a text cell cannot.

Banner discipline: where the CSV leads with banner lines (NOT-REPORTABLE
banners, preview disclaimers, simulated-data warnings, caveats), the XLSX
carries the SAME lines as its FIRST sheet ("Read first"), one line per row,
ahead of the data sheet — a spreadsheet user meets the caveats before the
numbers, exactly like a CSV reader does.

openpyxl is MIT (ADR-0001 tier 1 permissive; recorded in pyproject.toml and
checked by scripts/license_gate.py like every dependency).
"""

from __future__ import annotations

import csv
import io
import json
from typing import Iterable, Sequence

from fastapi import Response
from openpyxl import Workbook

#: The accepted values of every export endpoint's ``format`` query param.
FORMAT_PATTERN = "^(csv|xlsx)$"

#: Mirror of the web export's simulated cell (web/src/reports/csv.ts) — the
#: same words, so a simulated figure is equally unmissable in every format.
SIMULATED_CELL = "SIMULATED DATA - MUST NOT BE SUBMITTED"

#: Mirror of web/src/copy.ts copy.report.disclaimer — the Monthly Ridership
#: preview disclaimer the existing client-side CSV leads with.
METRICS_PREVIEW_DISCLAIMER = (
    "Preview only. The official NTD Monthly Ridership submission format has "
    "not yet been verified against FTA's reporting system documentation."
)

SIMULATED_BANNER = (
    "One or more rows below were computed from SIMULATED source data and "
    "are labeled in the simulated_data column. Simulated figures must never "
    "be submitted."
)

BANNER_SHEET_TITLE = "Read first"
DATA_SHEET_TITLE = "Data"

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
CSV_MEDIA_TYPE = "text/csv; charset=utf-8"


class Grid:
    """One assembled export: banner lines + header + all-string rows."""

    def __init__(
        self,
        banner_lines: Sequence[str],
        header: Sequence[str],
        rows: Iterable[Sequence[str]],
    ):
        self.banner_lines = [str(line) for line in banner_lines]
        self.header = [str(h) for h in header]
        self.rows = [[self._cell(c) for c in row] for row in rows]
        for row in self.rows:
            if len(row) != len(self.header):
                raise ValueError(
                    f"Export row has {len(row)} cells but the header has "
                    f"{len(self.header)} — a ragged export would misalign "
                    f"values under the wrong columns."
                )

    @staticmethod
    def _cell(value) -> str:
        """Cells must already be strings (or None for an empty cell) —
        anything else is refused loudly rather than coerced, because
        coercion is where a figure could silently change."""
        if value is None:
            return ""
        if not isinstance(value, str):
            raise TypeError(
                f"Export cells must be strings (figures exactly as served); "
                f"got {type(value).__name__}: {value!r}"
            )
        return value


def grid_to_csv(grid: Grid) -> bytes:
    """CSV: banner lines first (one single-cell line each), then header,
    then rows. CRLF line endings and minimal quoting — the web export's
    conventions (web/src/reports/csv.ts)."""
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    for line in grid.banner_lines:
        writer.writerow([line])
    writer.writerow(grid.header)
    for row in grid.rows:
        writer.writerow(row)
    return out.getvalue().encode("utf-8")


def grid_to_xlsx(grid: Grid) -> bytes:
    """XLSX: banner lines as the FIRST sheet (when any), data second. EVERY
    cell is a text cell — figures stay the exact strings the API serves."""
    wb = Workbook()
    data_sheet = wb.active
    if grid.banner_lines:
        data_sheet.title = BANNER_SHEET_TITLE
        for line in grid.banner_lines:
            data_sheet.append([line])
        data_sheet = wb.create_sheet(DATA_SHEET_TITLE)
    else:
        data_sheet.title = DATA_SHEET_TITLE
    data_sheet.append(grid.header)
    for row in grid.rows:
        # Empty string cells append as None in openpyxl; both read back as
        # "no content" — the CSV equality test normalizes both to "".
        data_sheet.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_response(grid: Grid, fmt: str, filename_stem: str) -> Response:
    """The one download response builder: same grid, either format, an
    attachment filename naming the surface and period."""
    if fmt == "csv":
        return Response(
            content=grid_to_csv(grid),
            media_type=CSV_MEDIA_TYPE,
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{filename_stem}.csv"'
                )
            },
        )
    return Response(
        content=grid_to_xlsx(grid),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_stem}.xlsx"'
            )
        },
    )


def is_simulated_detail(detail: dict) -> bool:
    """Mirror of web/src/detail.ts isSimulated: any source_mix key naming a
    simulated source marks the figure."""
    mix = detail.get("source_mix") if isinstance(detail, dict) else None
    if not isinstance(mix, dict):
        return False
    return any("simulated" in str(source).lower() for source in mix)


# ---------------------------------------------------------------------------
# Grid assemblies — one per export surface. Inputs are the EXACT objects the
# existing JSON surfaces serve (MetricValue models, the mr20/ss50 packages,
# the sampling progress payload); nothing is re-queried or re-derived here.
# ---------------------------------------------------------------------------

METRICS_VALUES_HEADER = (
    "metric",
    "unit",
    "period_start",
    "period_end",
    "scope",
    "value",
    "calc_name",
    "calc_version",
    "certification_status",
    "category",
    "simulated_data",
    "metric_value_id",
)


def metrics_values_grid(values) -> Grid:
    """GET /metrics/values as a grid: the web export's columns (web/src/
    reports/csv.ts CSV_HEADER) plus scope, category (the migration-0024
    honesty boundary) and metric_value_id (the provenance path — every
    exported figure names the row 'explain this number' starts from)."""
    any_simulated = any(is_simulated_detail(v.detail) for v in values)
    banner = [METRICS_PREVIEW_DISCLAIMER]
    if any_simulated:
        banner.append(SIMULATED_BANNER)
    rows = [
        [
            v.metric,
            v.unit,
            v.period_start.isoformat(),
            v.period_end.isoformat(),
            v.scope,
            v.value,  # VERBATIM — the exact string the API serves
            v.calc_name,
            v.calc_version,
            v.certification_status,
            v.category,
            SIMULATED_CELL if is_simulated_detail(v.detail) else "no",
            v.metric_value_id,
        ]
        for v in values
    ]
    return Grid(banner, METRICS_VALUES_HEADER, rows)


MR20_HEADER = (
    "scope",
    "metric",
    "value",
    "unit",
    "calc_name",
    "calc_version",
    "certification_status",
    "flags",
    "coverage",
    "non_reportable_pending_d2",
    "metric_value_id",
    "missing_reason",
)


def _mr20_rows_for_scope(scope_label: str, cells: dict, pending_d2: str):
    for metric in ("upt", "vrh", "vrm", "voms"):
        cell = cells.get(metric)
        if cell is None:
            continue
        if cell.get("value") is None:
            yield [
                scope_label, metric, "", "", "", "", "", "", "",
                pending_d2, "", cell.get("reason", ""),
            ]
        else:
            yield [
                scope_label,
                metric,
                cell["value"],  # VERBATIM package string
                cell.get("unit", ""),
                cell.get("calc_name", ""),
                cell.get("calc_version", ""),
                cell.get("certification_status", ""),
                "; ".join(cell.get("flags", [])),
                cell.get("coverage") or "",
                pending_d2,
                cell.get("metric_value_id", ""),
                "",
            ]


def mr20_grid(package: dict, certificate_lines: Sequence[str] = ()) -> Grid:
    """The headway_calc.mr20 package as a grid: its NOT-REPORTABLE banner
    and every programmatically enumerated caveat lead (banner lines / first
    sheet), then one row per (scope, metric) cell, values verbatim.

    ``certificate_lines`` (handoff 0019, design point 7): when the period
    holds certified figures, the caller passes the certificate block
    (signer, timestamp, key fingerprint per certification —
    routers/reports._certificate_lines) and it joins the banner / "Read
    first" sheet after the caveats. Empty when nothing is certified —
    no line is ever invented."""
    banner = [
        package["banner"],
        f"Form {package['form']} preview for {package['month']} "
        f"(period [{package['period_start']}, {package['period_end']}), "
        f"{package['period_convention']}).",
        f"Citation: {package['citation']}",
        f"Generated by {package['generator']['name']} "
        f"{package['generator']['version']}.",
    ]
    banner.extend(
        f"Caveat [{c['id']}, {c['status']}]: {c['text']}"
        for c in package["caveats"]
    )
    banner.extend(str(line) for line in certificate_lines)
    rows = list(_mr20_rows_for_scope("fleet", package["fleet"], ""))
    for mode in sorted(package["modes"]):
        entry = package["modes"][mode]
        pending = "yes" if entry.get("non_reportable_pending_d2") else "no"
        rows.extend(_mr20_rows_for_scope(f"mode:{mode}", entry, pending))
    return Grid(banner, MR20_HEADER, rows)


SS50_HEADER = (
    "mode",
    "type_of_service",
    "zero_event",
    "injury_events",
    "people_injured",
    "injury_event_ids",
    "non_major_fires",
    "fire_event_ids",
    "assaults_on_worker",
    "assaults_without_injury",
    "assault_event_ids",
)


def ss50_grid(package: dict, certificate_lines: Sequence[str] = ()) -> Grid:
    """The headway_calc.ss50 package as a grid: banner + citations + caveats
    + the excluded-event accounting lead; then one row per (mode, TOS) cell
    including the explicit zero rows.

    ``certificate_lines`` exactly as mr20_grid: the period's certificate
    block joins the banner / "Read first" sheet when certified figures
    exist for the month; empty otherwise."""
    banner = [
        package["banner"],
        f"Form {package['form']} preview for {package['month']} "
        f"(period [{package['period_start']}, {package['period_end']}), "
        f"{package['period_convention']}). Due date: {package['due_date']}.",
        f"Generated by {package['generator']['name']} "
        f"{package['generator']['version']}.",
    ]
    banner.extend(
        f"Citation [{c['id']}]: {c['text']} ({c['source']})"
        for c in package["citations"]
    )
    banner.extend(
        f"Caveat [{c['id']}]: {c['text']}" for c in package["caveats"]
    )
    banner.extend(str(line) for line in certificate_lines)
    excluded = package["excluded"]
    for label, key in (
        ("major (S&S-40, not counted here)", "major_event_ids"),
        ("not reportable", "not_reportable_event_ids"),
        ("superseded (replacement carries the truth)", "superseded_event_ids"),
        ("unclassified (classify and regenerate)", "unclassified_event_ids"),
    ):
        ids = excluded.get(key, [])
        if ids:
            banner.append(
                f"Excluded events — {label}: {', '.join(ids)}"
            )
    rows = []
    for cell in package["cells"]:
        counts = cell["counts"]
        rows.append(
            [
                cell["mode"],
                cell["type_of_service"],
                "yes" if cell["zero_event"] else "no",
                str(counts["injury_events"]["count"]),
                str(counts["injury_events"]["people_injured"]),
                "; ".join(counts["injury_events"]["event_ids"]),
                str(counts["non_major_fires"]["count"]),
                "; ".join(counts["non_major_fires"]["event_ids"]),
                str(counts["assaults_on_worker"]["count"]),
                str(counts["assaults_on_worker"]["without_injury"]),
                "; ".join(counts["assaults_on_worker"]["event_ids"]),
            ]
        )
    return Grid(banner, SS50_HEADER, rows)


SAMPLING_WORKSHEET_HEADER = (
    "period_label",
    "unit_id",
    "measured",
    "draw_oversample_units",
)


def sampling_worksheet_grid(
    plan, draws: list[dict], measured_unit_ids: set[str], retention_note: str
) -> Grid:
    """The sampling plan's measurement worksheet as a grid: one row per
    selected unit per draw (the same draw records GET /sampling/plans/{id}/
    progress summarizes), measured state from the plan's ACTIVE
    measurements; the plan's requirement, the undersampled/estimate-ready
    state and the retention note lead as banner lines."""
    measured_overall = len(measured_unit_ids)
    undersampled = measured_overall < plan.required_annual
    banner = [
        f"Sampling worksheet — plan {plan.plan_id}: report year "
        f"{plan.report_year}, mode {plan.mode}, TOS {plan.type_of_service}, "
        f"unit {plan.unit}, {plan.efficiency_option} option, "
        f"{plan.frequency} draws.",
        f"Required: {plan.required_per_period} per period, "
        f"{plan.required_annual} for the year ({plan.table_citation})",
        f"Measured so far: {measured_overall} of "
        f"{plan.required_annual} required — "
        + (
            "UNDERSAMPLED: the estimate is refused until every required "
            "unit is measured."
            if undersampled
            else "requirement met; the plan is estimate-ready."
        ),
        retention_note,
    ]
    rows = []
    for draw in draws:
        for unit in draw["selected_units"]:
            rows.append(
                [
                    draw["period_label"],
                    unit,
                    "yes" if unit in measured_unit_ids else "no",
                    str(draw["oversample_units"]),
                ]
            )
    return Grid(banner, SAMPLING_WORKSHEET_HEADER, rows)


# ---------------------------------------------------------------------------
# Multi-sheet workbooks (handoff 0020) — the same one-assembly invariant,
# extended: one banner-line list + N titled sheets feed BOTH formats. In the
# CSV, each sheet is introduced by ONE single-cell marker line
# (SHEET_MARKER_PREFIX + title) followed by its header and rows; in the XLSX
# the marker line becomes the sheet's TAB NAME (never a data row), the
# banner lines are the first ("Read first") sheet, and every cell stays a
# TEXT cell holding the byte-identical string the CSV holds.
# ---------------------------------------------------------------------------

SHEET_MARKER_PREFIX = "## "


class SheetGrid:
    """One titled sheet of a multi-sheet export: a Grid without banner
    lines (the workbook's banner is workbook-level, not per-sheet)."""

    def __init__(self, title: str, header: Sequence[str], rows):
        if not title or len(title) > 31:
            # Excel refuses sheet titles over 31 chars; refuse loudly here
            # rather than let openpyxl mangle a name the CSV spells out.
            raise ValueError(f"Sheet title must be 1..31 chars: {title!r}")
        self.title = title
        self.grid = Grid([], header, rows)


def sheets_to_csv(banner_lines: Sequence[str], sheets: Sequence[SheetGrid]) -> bytes:
    """CSV: banner lines first (one single-cell line each), then per sheet a
    single-cell marker line ('## <title>'), its header, and its rows."""
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    for line in banner_lines:
        writer.writerow([str(line)])
    for sheet in sheets:
        writer.writerow([f"{SHEET_MARKER_PREFIX}{sheet.title}"])
        writer.writerow(sheet.grid.header)
        for row in sheet.grid.rows:
            writer.writerow(row)
    return out.getvalue().encode("utf-8")


def sheets_to_xlsx(banner_lines: Sequence[str], sheets: Sequence[SheetGrid]) -> bytes:
    """XLSX: the banner lines as the FIRST sheet ('Read first' — a
    spreadsheet user meets the caveats before the numbers), then one sheet
    per SheetGrid, tab-named by its title. Text cells only."""
    if not banner_lines:
        raise ValueError(
            "A multi-sheet workbook must lead with banner lines — the "
            "'Read first' sheet is not optional."
        )
    wb = Workbook()
    banner_sheet = wb.active
    banner_sheet.title = BANNER_SHEET_TITLE
    for line in banner_lines:
        banner_sheet.append([str(line)])
    for sheet in sheets:
        ws = wb.create_sheet(sheet.title)
        ws.append(sheet.grid.header)
        for row in sheet.grid.rows:
            ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def sheets_export_response(
    banner_lines: Sequence[str],
    sheets: Sequence[SheetGrid],
    fmt: str,
    filename_stem: str,
) -> Response:
    """The multi-sheet download response builder: same assembly, either
    format (the export_response sibling)."""
    if fmt == "csv":
        return Response(
            content=sheets_to_csv(banner_lines, sheets),
            media_type=CSV_MEDIA_TYPE,
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{filename_stem}.csv"'
                )
            },
        )
    return Response(
        content=sheets_to_xlsx(banner_lines, sheets),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_stem}.xlsx"'
            )
        },
    )


# ---------------------------------------------------------------------------
# The monthly agency workbook (handoff 0020, design point 3) — OUR OWN
# generic layout: "Read first" banner sheet, a Ridership-by-mode sheet and
# an Operations sheet, every data cell either a VERBATIM served figure with
# a visible provenance column (metric_value_id — we're proud of it) or an
# EXPLICIT stated absence. This module formats; it never computes: no
# arithmetic, no rounding, no derived deltas (year-over-year columns show
# the prior year month's persisted figure verbatim with its OWN provenance;
# comparison arithmetic lives in GET /metrics/compare, the one blessed
# comparison affordance).
# ---------------------------------------------------------------------------

WORKBOOK_GENERATOR_NAME = "headway_api.exports.agency_workbook"
WORKBOOK_GENERATOR_VERSION = "0.1.0"

WORKBOOK_SHEET_RIDERSHIP = "Ridership by mode"
WORKBOOK_SHEET_OPERATIONS = "Operations"

WORKBOOK_HEADER = (
    "measure",
    "scope",
    "value",
    "unit",
    "prior_year_value",
    "prior_year_provenance",
    "calc_name",
    "calc_version",
    "certification_status",
    "category",
    "simulated_data",
    "provenance_metric_value_id",
    "note",
)

#: The stated-absence cell (binding rule: absent cells STATED, never
#: invented, never zero-filled).
WORKBOOK_ABSENT_CELL = "no figure for this period — not yet computed by Headway"
WORKBOOK_PRIOR_ABSENT_CELL = "no figure for this period"

WORKBOOK_BANNER = (
    "NOT REPORTABLE — automatically assembled monthly workbook preview. "
    "Every figure below is served VERBATIM from computed.metric_values "
    "with its provenance id; nothing in this workbook may be submitted to "
    "the NTD."
)

#: The migration-0024 wall, stated where the ops rows live.
WORKBOOK_OPS_BADGE = (
    "OPERATIONS METRIC — never certifiable (database-enforced, migration "
    "0024); not an NTD figure; never part of any submission."
)

#: Honest scope (handoff 0020, design point 4): rows the workbook STATES it
#: does not compute, so their absence is a fact, not a gap.
WORKBOOK_MISSED_TRIPS_NOTE = (
    "Not computed by Headway v0: missed-trip accounting needs "
    "schedule-vs-operated reconciliation semantics that have not been "
    "verified against the NTD manuals yet — a known future increment "
    "(handoff 0020). Stated absent, never zero-filled."
)
WORKBOOK_VAMS_NOTE = (
    "Not computed by Headway: VAMS (Vehicles Available for Maximum Service) "
    "requires fleet inventory data Headway does not ingest yet. Stated "
    "absent, never zero-filled."
)
WORKBOOK_DAYS_NOT_OPERATED_NOTE = (
    "Not computed by Headway v0: Days Not Operated Due to Strikes / "
    "Officially Declared Emergencies (2026 NTD Policy Manual p. 155) need "
    "agency declarations the settings surface does not capture yet. Stated "
    "absent, never zero-filled."
)

_DAY_TYPES = ("weekday", "saturday", "sunday")


def _detail_dict(detail) -> dict:
    """detail arrives as a dict (psycopg JSONB) or JSON text (fakes/other
    drivers); anything else is refused loudly (the mr20 convention)."""
    if detail is None:
        return {}
    if isinstance(detail, dict):
        return detail
    if isinstance(detail, str):
        return json.loads(detail)
    raise TypeError(f"Unsupported detail payload type: {type(detail).__name__}")


class _WorkbookCells:
    """Index of one period's latest rows by (metric, scope) — tuples in the
    reports._SELECT_WORKBOOK_LATEST column order."""

    def __init__(self, rows):
        self.by_key = {}
        for row in rows:
            (metric, scope, metric_value_id, value, unit, calc_name,
             calc_version, certification_status, category, detail) = row
            self.by_key[(metric, scope)] = {
                "metric_value_id": str(metric_value_id),
                "value": str(value),
                "unit": unit,
                "calc_name": calc_name,
                "calc_version": calc_version,
                "certification_status": certification_status,
                "category": category,
                "detail": _detail_dict(detail),
            }

    def get(self, metric: str, scope: str):
        return self.by_key.get((metric, scope))

    def plain_mode_scopes(self) -> list[str]:
        """The plain 'mode:<mode>' scopes present (exactly two segments —
        excludes 'mode:DR:tos:*' and 'mode:<m>:daytype:*' refinements)."""
        return sorted(
            {
                scope
                for (_metric, scope) in self.by_key
                if scope.startswith("mode:") and scope.count(":") == 1
            }
        )

    def any_simulated(self) -> bool:
        return any(
            is_simulated_detail(cell["detail"])
            for cell in self.by_key.values()
        )


def _workbook_row(
    measure: str,
    scope_label: str,
    cell,
    prior_cell,
    note: str = "",
) -> list[str]:
    """One workbook row: the cell verbatim, or the stated absence."""
    if cell is None:
        value, unit = WORKBOOK_ABSENT_CELL, ""
        calc_name = calc_version = certification = category = ""
        simulated = ""
        provenance = ""
        if note:
            note = note
        else:
            note = (
                "No computed.metric_values row for this measure and scope "
                "in the period: not computed yet, or its run refused and "
                "routed the reason to dq.issues. A missing figure is "
                "reported as missing, never invented."
            )
    else:
        value = cell["value"]  # VERBATIM — the exact string the API serves
        unit = cell["unit"]
        calc_name = cell["calc_name"]
        calc_version = cell["calc_version"]
        certification = cell["certification_status"]
        category = cell["category"]
        simulated = (
            SIMULATED_CELL if is_simulated_detail(cell["detail"]) else "no"
        )
        provenance = cell["metric_value_id"]
    if prior_cell is None:
        prior_value = WORKBOOK_PRIOR_ABSENT_CELL
        prior_provenance = ""
    else:
        prior_value = prior_cell["value"]
        prior_provenance = prior_cell["metric_value_id"]
    return [
        measure,
        scope_label,
        value,
        unit,
        prior_value,
        prior_provenance,
        calc_name,
        calc_version,
        certification,
        category,
        simulated,
        provenance,
        note,
    ]


def _avg_note(cell) -> str:
    """The typical/atypical statement a day-type average row carries, built
    from FACTS in the persisted detail (formatting, never arithmetic)."""
    if cell is None:
        return ""
    detail = cell["detail"]
    days = detail.get("days_operated")
    split = detail.get("split", "typical")
    note = f"average over {days} operated {detail.get('day_type', '')} day(s), {split} split"
    if detail.get("atypical_flags_declared") is False:
        note += (
            "; no atypical days declared for this period — every day is "
            "typical (stated, per the agency's service-day declarations)"
        )
    return note


def _days_operated_note(cell) -> str:
    if cell is None:
        return ""
    detail = cell["detail"]
    unobserved = detail.get("unobserved_dates", [])
    if unobserved:
        return (
            f"observed lower bound: {len(unobserved)} date(s) of this "
            f"schedule type had no telemetry (see the "
            f"daytype_days_unobserved finding)"
        )
    return "every date of this schedule type was observed"


def agency_workbook(
    month: str,
    period_start,
    period_end,
    prior_month: str,
    rows,
    prior_rows,
    certificate_lines: Sequence[str] = (),
) -> tuple[list[str], list[SheetGrid]]:
    """Assemble the monthly agency workbook (banner lines + two sheets) from
    the latest persisted rows of the month and of the prior-year month.

    OUR OWN generic layout (never a partner file's): a "Read first" banner,
    a Ridership-by-mode sheet (UPT month totals per mode, day-type averages
    with typical/atypical splits, Days Operated per schedule type, the
    stated-absent missed-trip and days-not-operated rows) and an Operations
    sheet (VOMS per mode — an NTD figure — plus the migration-0024-badged
    OTP/headway-adherence rows and the stated-absent VAMS row). Every data
    cell is the VERBATIM served string; every present cell names its
    metric_value_id in the visible provenance column; every absent cell is
    STATED. Year-over-year columns show the prior year month's persisted
    figure verbatim with its own provenance — never a derived delta (this
    module never computes; GET /metrics/compare is the comparison surface).
    """
    cells = _WorkbookCells(rows)
    prior = _WorkbookCells(prior_rows)

    banner = [
        WORKBOOK_BANNER,
        f"Monthly agency workbook for {month} (period "
        f"[{period_start.isoformat()}, {period_end.isoformat()}), half-open, "
        f"UTC).",
        f"Generated by {WORKBOOK_GENERATOR_NAME} "
        f"{WORKBOOK_GENERATOR_VERSION}.",
        "How to read this workbook: every data cell is the exact figure "
        "string Headway serves (text cells — a figure is never coerced "
        "through a floating-point number), and every present figure names "
        "the computed.metric_values row it came from in the "
        "provenance_metric_value_id column — 'explain this number' starts "
        "there (GET /metrics/values/{id}/lineage).",
        "Absent cells are STATED ('" + WORKBOOK_ABSENT_CELL + "') — "
        "Headway never invents or zero-fills a cell it did not compute.",
        f"Year-over-year columns show the prior year month's ({prior_month}) "
        f"persisted figure verbatim with its own provenance id; Headway "
        f"deliberately does not derive a delta in an export — comparison "
        f"arithmetic lives in GET /metrics/compare (exact-decimal, "
        f"sign-neutral).",
        "Day-type figures (average weekday/Saturday/Sunday UPT, Days "
        "Operated per schedule) are classified by calc daytype_v0 0.1.0: "
        "agency-declared overrides in app.service_day_overrides govern "
        "(holiday reassignments per 2026 NTD Policy Manual p. 156; "
        "atypical-day flags); otherwise day-of-week. Basis quoted in "
        "services/calc/REGULATORY_TRACKER.md, 'Verified — Days Operated "
        "and day-type schedules'.",
        "Typical/atypical splits appear only where the agency declared "
        "atypical days; an unflagged month is all typical — stated on each "
        "average row, never silently assumed.",
        "Known absences (honest scope, handoff 0020): "
        + WORKBOOK_MISSED_TRIPS_NOTE
        + " "
        + WORKBOOK_VAMS_NOTE,
        "The Operations sheet mixes NTD figures (VOMS) with OPERATIONS "
        "metrics (on-time performance, headway adherence). "
        + WORKBOOK_OPS_BADGE,
    ]
    if cells.any_simulated() or prior.any_simulated():
        banner.append(SIMULATED_BANNER)
    banner.extend(str(line) for line in certificate_lines)

    # --- Ridership by mode ---------------------------------------------------
    ridership: list[list[str]] = []
    mode_scopes = cells.plain_mode_scopes()
    for scope in ["agency"] + mode_scopes:
        label = "all modes (agency)" if scope == "agency" else scope
        ridership.append(
            _workbook_row(
                "Unlinked Passenger Trips (month total)",
                label,
                cells.get("upt", scope),
                prior.get("upt", scope),
            )
        )
    for day_type in _DAY_TYPES:
        scope = f"daytype:{day_type}"
        cell = cells.get("upt_avg", scope)
        ridership.append(
            _workbook_row(
                f"Average {day_type} UPT (typical days)",
                "all modes (agency)",
                cell,
                prior.get("upt_avg", scope),
                note=_avg_note(cell),
            )
        )
        atypical_cell = cells.get("upt_avg", f"{scope}:atypical")
        if atypical_cell is not None:
            ridership.append(
                _workbook_row(
                    f"Average {day_type} UPT (agency-declared atypical days)",
                    "all modes (agency)",
                    atypical_cell,
                    prior.get("upt_avg", f"{scope}:atypical"),
                    note=_avg_note(atypical_cell),
                )
            )
    for mode_scope in mode_scopes:
        for day_type in _DAY_TYPES:
            scope = f"{mode_scope}:daytype:{day_type}"
            cell = cells.get("upt_avg", scope)
            ridership.append(
                _workbook_row(
                    f"Average {day_type} UPT (typical days)",
                    mode_scope,
                    cell,
                    prior.get("upt_avg", scope),
                    note=_avg_note(cell),
                )
            )
            atypical_cell = cells.get("upt_avg", f"{scope}:atypical")
            if atypical_cell is not None:
                ridership.append(
                    _workbook_row(
                        f"Average {day_type} UPT (agency-declared atypical "
                        f"days)",
                        mode_scope,
                        atypical_cell,
                        prior.get("upt_avg", f"{scope}:atypical"),
                        note=_avg_note(atypical_cell),
                    )
                )
    for day_type in _DAY_TYPES:
        scope = f"daytype:{day_type}"
        cell = cells.get("days_operated", scope)
        ridership.append(
            _workbook_row(
                f"Days operated ({day_type} schedule)",
                "all modes (agency)",
                cell,
                prior.get("days_operated", scope),
                note=_days_operated_note(cell),
            )
        )
    ridership.append(
        _workbook_row(
            "Days not operated (strikes / declared emergencies)",
            "all modes (agency)",
            None,
            None,
            note=WORKBOOK_DAYS_NOT_OPERATED_NOTE,
        )
    )
    ridership.append(
        _workbook_row(
            "Missed trips / trips not operated",
            "all modes (agency)",
            None,
            None,
            note=WORKBOOK_MISSED_TRIPS_NOTE,
        )
    )

    # --- Operations ----------------------------------------------------------
    operations: list[list[str]] = []
    for scope in ["agency"] + mode_scopes:
        label = "all modes (agency)" if scope == "agency" else scope
        operations.append(
            _workbook_row(
                "Vehicles Operated in Maximum Service (VOMS)",
                label,
                cells.get("voms", scope),
                prior.get("voms", scope),
            )
        )
    operations.append(
        _workbook_row(
            "Vehicles Available for Maximum Service (VAMS)",
            "all modes (agency)",
            None,
            None,
            note=WORKBOOK_VAMS_NOTE,
        )
    )
    for metric, measure in (
        ("otp", "On-time performance (share of observed passages on time)"),
        (
            "headway_adherence",
            "Headway adherence (coefficient of variation of headways)",
        ),
    ):
        cell = cells.get(metric, "agency")
        operations.append(
            _workbook_row(
                measure,
                "all modes (agency)",
                cell,
                prior.get(metric, "agency"),
                note=WORKBOOK_OPS_BADGE,
            )
        )

    sheets = [
        SheetGrid(WORKBOOK_SHEET_RIDERSHIP, WORKBOOK_HEADER, ridership),
        SheetGrid(WORKBOOK_SHEET_OPERATIONS, WORKBOOK_HEADER, operations),
    ]
    return banner, sheets
